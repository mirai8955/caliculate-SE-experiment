#########################
##集計結果は基本(2,2)に設定する
##stiListに提示種類を追加する
#########################
import openpyxl
import numpy as np
import math

wb = openpyxl.load_workbook("C:\\Users\\okalaptop\\Documents\\研究\\lifetech\\実験結果.xlsx")
ws = wb["Sheet1"]
ws2 = wb.create_sheet(title = "Result")




def total(ws,ws2,stiCol, ansCol, setRow, setCol, trialNum, stiList):
    resultArray = np.zeros((len(stiList),len(stiList)))

    ws2.cell(setRow,setCol).value =  ws.cell(1,ansCol).value 
    for i in range(1,len(stiList)+1):
        ws2.cell(setRow, setCol+i).value = stiList[i-1]
    for i in range(1,len(stiList)+1):
        ws2.cell(setRow+i, setCol).value = stiList[i-1]
    for i in range(2, trialNum + 2):
        j = 0
        k = 0
        while(1):
            if str(ws.cell(i,stiCol).value) == stiList[j]:
                break
            j = j+1
        while(1):
            if str(ws.cell(i,ansCol).value) == stiList[k]:
                break
            k = k+1
        resultArray[j][k] = resultArray[j][k] + 0.1

    for i in range(len(stiList)):
        for j in range(len(stiList)):
            ws2.cell(setRow+1+i,setCol+1+j).value = resultArray[i][j]


def computeAve1(ws2,setRow,setCol,number,stiList,stiList2):
    stiKind = len(stiList2)
    averageArray = np.zeros((stiKind, stiKind))

    ws2.cell(setRow,setCol).value =  "正答率"
    for i in range(1,len(stiList2)+1):
        ws2.cell(setRow, setCol+i).value = stiList2[i-1]
    for i in range(1,len(stiList2)+1):
        ws2.cell(setRow+i, setCol).value = stiList2[i-1]

    #平均を計算
    for o in range(2):
        for p in range(2):
            for i in range(3):
                for j in range(number):
                    row = 3+o*3+(len(stiList)+4)*j+i
                    col = 3+p*3
                    averageArray[o][p] =  averageArray[o][p] + ws2.cell(row, col).value + ws2.cell(row, col).value + ws2.cell(row, col).value
                    if i == 0 : print(averageArray[o][p])
            averageArray[o][p] = averageArray[o][p]/number
            averageArray[o][p] = averageArray[o][p]/3
            ws2.cell(setRow+1+o,setCol+1+p).value = averageArray[o][p]
            #print(averageArray)

def computeAve2(ws2, setRow, setCol, number, stiList):
    averageArray = np.zeros((len(stiList),len(stiList)))

    ws2.cell(setRow,setCol).value =  "正答率"
    for i in range(1,len(stiList)+1):
        ws2.cell(setRow, setCol+i).value = stiList[i-1]
    for i in range(1,len(stiList)+1):
        ws2.cell(setRow+i, setCol).value = stiList[i-1]
    
    #平均を計算
    for i in range(len(stiList)):
        for j in range(len(stiList)):
            for k in range(number):
                averageArray[i][j] =  averageArray[i][j] + ws2.cell(3+(len(stiList)+4)*k+i,3+j).value
            averageArray[i][j] = averageArray[i][j]/number
            ws2.cell(setRow+1+i,setCol+1+j).value = averageArray[i][j]



def computeSE1(ws2,setRow,setCol,number,stiList,stiList2):
    stiKind = len(stiList2)
    averageArray = np.zeros((stiKind, stiKind))
    SEArray = np.zeros((stiKind, stiKind))

    ws2.cell(setRow,setCol).value =  "SE"
    for i in range(1,len(stiList2)+1):
        ws2.cell(setRow, setCol+i).value = stiList2[i-1]
    for i in range(1,len(stiList2)+1):
        ws2.cell(setRow+i, setCol).value = stiList2[i-1]

    #平均を計算
    for o in range(2):
        for p in range(2):
            for i in range(3):
                for j in range(number):
                    averageArray[o][p] =  averageArray[o][p] + ws2.cell(3+o*3+(len(stiList)+4)*j+i,3+p*3).value + ws2.cell(3+o*3+(len(stiList)+4)*j+i,4+p*3).value + ws2.cell(3+o*3+(len(stiList)+4)*j+i,5+p*3).value
    
    averageArray = averageArray/number
    averageArray = averageArray/3

    #SEを計算
    sum = 0
    for o in range(2):
        for p in range(2):
            for i in range(number):
                sum = 0
                for j in range(3):
                    sum =  sum + ws2.cell(3+o*3+(len(stiList)+4)*i+j,3+p*3).value + ws2.cell(3+o*3+(len(stiList)+4)*i+j,4+p*3).value + ws2.cell(3+o*3+(len(stiList)+4)*i+j,5+p*3).value
                sum = sum/3
                SEArray[o][p] = SEArray[o][p] + (sum - averageArray[o][p])**2
            SEArray[o][p] = math.sqrt(SEArray[o][p]/(number-1))/math.sqrt(number)
            ws2.cell(setRow+1+o,setCol+1+p).value = SEArray[o][p]

    
    


def computeSE2(ws2, setRow, setCol, number, stiList):
    averageArray = np.zeros((len(stiList),len(stiList)))
    SEArray = np.zeros((len(stiList),len(stiList)))

    ws2.cell(setRow,setCol).value =  "SE"
    for i in range(1,len(stiList)+1):
        ws2.cell(setRow, setCol+i).value = stiList[i-1]
    for i in range(1,len(stiList)+1):
        ws2.cell(setRow+i, setCol).value = stiList[i-1]
    
    #平均を計算
    for i in range(len(stiList)):
        for j in range(len(stiList)):
            for k in range(number):
                averageArray[i][j] =  averageArray[i][j] + ws2.cell(3+(len(stiList)+4)*k+i,3+j).value
    averageArray = averageArray/number

    #SEを計算
    for i in range(len(stiList)):
        for j in range(len(stiList)):
            for k in range(number):
                SEArray[i][j] = SEArray[i][j] + (ws2.cell(3+(len(stiList)+4)*k+i,3+j).value - averageArray[i][j])**2
            SEArray[i][j] = math.sqrt(SEArray[i][j]/(number-1))/math.sqrt(number)
            ws2.cell(setRow+1+i,setCol+1+j).value = SEArray[i][j]
    
    
##以下から記述

stiList = ['1', '2', '3', 'a', 'b', 'c']
stiList2 = ['bump', 'dent']
number = 5
setRow = 2
setCol = 2
for i in range(number):
    total(ws, ws2, 27, 28+i, setRow+(len(stiList)+4)*i, setCol, 60, stiList)

computeAve1(ws2, 2, 11, number, stiList, stiList2)
computeSE1(ws2, 7, 11, number, stiList, stiList2)
computeAve2(ws2, 12, 11, number, stiList)
computeSE2(ws2, 22, 11, number, stiList)
      


wb.save("ResultsOfExperiment.xlsx")
wb.close()